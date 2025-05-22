import streamlit as st 
import openpyxl
from openpyxl.styles import Font
from io import BytesIO
import base64

# ---------------- Page Config ---------------- #
st.set_page_config(page_title="noqari 1.0", layout="centered")

# ---------------- Custom CSS ---------------- #
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Lexend:wght@400;700&family=DM+Serif+Display&display=swap');

/* Base page */
html, body, [class*="css"] {
  font-family: 'Lexend', sans-serif;
  background-color: #ffffff;
  padding: 24px;
}

/* Card container */
.block-container, section.main {
  background-color: #ffffff !important;
  border-radius: 18px;
  padding: 3rem 2rem;
  box-shadow: 0 8px 24px rgba(0,0,0,0.08);
  max-width: 800px;
  margin: auto;
}

/* Title */
.title-text {
  font-family: 'Georgia', serif;
  font-size: 3rem;
  font-weight: bold;
  color: #111111;
  text-align: center;
  margin-bottom: 0.2rem;
}

/* Tagline */
.tagline {
  font-family: 'DM Serif Display', serif;
  text-align: center;
  font-size: 1.4rem;
  margin: 0.5rem 0 1.5rem;
  color: #FF69B4;
}

/* Uploader box */
.uploadbox {
  padding: 1rem;
  border-radius: 12px;
  background-color: #ffffff;
  border: 1px solid #e6e6e6;
  box-shadow: 0 2px 8px rgba(0,0,0,0.05);
  margin-bottom: 1.5rem;
}

/* Hide default uploader label */
section[data-testid="stFileUploader"] label {
  display: none !important;
}

/* Center info alert */
div[data-testid="stAlert"] {
  text-align: center;
}

/* Browse files button */
div[data-testid="stFileUploader"] button {
  background-color: #FF69B4 !important;
  color: #ffffff !important;
  border: none !important;
  position: relative;
  overflow: hidden;
}
div[data-testid="stFileUploader"] button::after {
  content: "";
  position: absolute; top: 0; left: -100%;
  width: 100%; height: 100%;
  background: linear-gradient(120deg,
    rgba(255,255,255,0.2),
    rgba(255,255,255,0.5),
    rgba(255,255,255,0.2)
  );
  transition: all 0.5s ease-in-out;
}
div[data-testid="stFileUploader"] button:hover::after {
  left: 100%;
}

/* Footer text */
.footer-note {
  font-size: 0.95rem;
  text-align: center;
  margin-top: 50px;
  color: #333333;
}
.thank-you {
  font-family: 'Georgia', serif;
  text-align: center;
  font-size: 1.1rem;
  color: #FF69B4;
  margin-top: 8px;
}
</style>
""", unsafe_allow_html=True)

# ---------------- Header & Tagline ---------------- #
st.markdown("""
<div class="title-text">noqari 1.0</div>
<div style="text-align:center; font-size:1.6rem;">💌</div>
<div class="tagline">the happiest place on earth (for VLOOKUP formulas).</div>
""", unsafe_allow_html=True)

# ---------------- File Upload ---------------- #
st.markdown('<div class="uploadbox">', unsafe_allow_html=True)
uploaded_file = st.file_uploader(
    "Upload your PCARD_OPEN.xlsx",
    type="xlsx",
    label_visibility="hidden"
)
st.markdown('</div>', unsafe_allow_html=True)

# ---------------- Info Message ---------------- #
st.markdown(
    '<div style="text-align:center; background-color:#eaf3fc; padding:1rem; '
    'border-radius:8px; margin-bottom:2rem;">'
    'Please upload your PCARD_OPEN.xlsx file to get started!'
    '</div>',
    unsafe_allow_html=True
)

# ---------------- Excel Logic (Pure-Values) ---------------- #
if uploaded_file:
    # hide default alert
    st.markdown("<div></div>", unsafe_allow_html=True)

    wb = openpyxl.load_workbook(uploaded_file)
    sheet1 = wb.worksheets[0]
    sheet2 = wb.worksheets[1]

    # Utility to find last data row by scanning a given column (e.g., F=6)
    def find_last_data_row(sheet, col_idx):
        for r in range(sheet.max_row, 1, -1):
            if sheet.cell(r, col_idx).value not in (None, ""):
                return r
        return 1

    # 1) A-column formulas in both sheets, capped at real data end
    for sht in (sheet1, sheet2):
        last = find_last_data_row(sht, 6)
        for r in range(2, last + 1):
            c = sht[f"A{r}"]
            c.value = f"=F{r}&G{r}&H{r}"
            c.font = Font(name="Calibri", size=11)

    # 2) Build lookup dict from Sheet1, dropping only truly blank data (not zero)
    lookup = {}
    for r in range(2, sheet1.max_row + 1):
        f = sheet1.cell(r, 6).value or ""
        g = sheet1.cell(r, 7).value or ""
        h = sheet1.cell(r, 8).value or ""
        key = f"{f}{g}{h}"
        p = sheet1.cell(r, 16).value
        q = sheet1.cell(r, 17).value
        lookup[key] = (
            p if p not in (None, "") else None,
            q if q not in (None, "") else None
        )

    # 3) Write static lookup values into Sheet2's P & Q, capped at last data row,
    #    and never write when there's nothing to pull
    last2 = find_last_data_row(sheet2, 1)  # anchor on column A of Sheet2
    for r in range(2, last2 + 1):
        f = sheet2.cell(r, 6).value or ""
        g = sheet2.cell(r, 7).value or ""
        h = sheet2.cell(r, 8).value or ""
        key = f"{f}{g}{h}"

        p_val, q_val = lookup.get(key, (None, None))

        if p_val is not None:
            sheet2.cell(r, 16).value = p_val
        if q_val is not None:
            sheet2.cell(r, 17).value = q_val

    # 4) Save & provide download link
    buf = BytesIO()
    wb.save(buf)
    b64 = base64.b64encode(buf.getvalue()).decode()

    st.markdown(
        "<div style='text-align:center; font-size:1.2rem; margin-top:1.5rem;'>"
        "✨ All yours! Your file is ready to go!! ✨</div>",
        unsafe_allow_html=True
    )
    st.markdown(f"""
    <div style="text-align:center; margin-top:2rem;">
      <a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}"
         download="PCARD_OPEN_Processed.xlsx"
         style="
           display: inline-block;
           padding: 0.75rem 1.5rem;
           font-size: 1rem;
           font-weight: 600;
           color: white;
           background-color: #FF69B4;
           border: none;
           border-radius: 10px;
           text-decoration: none;
           box-shadow: 0 4px 12px rgba(0,0,0,0.15);
           transition: all 0.3s ease-in-out;
         "
         onmouseover="this.style.opacity=0.9"
         onmouseout="this.style.opacity=1"
      >
        Download Processed File
      </a>
    </div>
    """, unsafe_allow_html=True)

# ---------------- Footer ---------------- #
st.markdown("""
<div class="footer-note">
  <strong>NOTE:</strong> To ensure the code runs correctly, the file must be renamed to 
  <code>PCARD_OPEN</code> and saved in <code>.xlsx</code> format.<br>
  Files with a different name or format will not be processed.
</div>
<div class="thank-you">sincerely, your tiny tab fairy</div>
""", unsafe_allow_html=True)
